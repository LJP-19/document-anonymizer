"""Client roster and the pseudonym mapping workbook.

Two jobs.

**Consistency.** The same client must get the same pseudonym in their 1040,
their K-1 and their bank statement, so a batch of documents stays internally
coherent for whoever analyses them downstream. The roster carries the mapping
between runs.

**Reversal.** Analysis comes back talking about the pseudonym. The workbook is
how you translate it back to the real client.

SECURITY: this workbook lists real client identifiers beside their pseudonyms.
It is the one file in this system that must never be sent anywhere. It is
written to a separate folder from the anonymised PDF and named so it cannot be
attached by accident.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from ..detection.types import PiiType

MAPPING_SUFFIX = "MAPPING-DO-NOT-SEND.xlsx"
SHEET_NAME = "Mapping"
ROSTER_FILENAME = "client-roster.xlsx"

#: Types worth carrying between documents. A pseudonymised city is not worth
#: pinning across a whole engagement; a name, an SSN and an EIN are.
ROSTER_TYPES = {
    PiiType.PERSON, PiiType.SSN, PiiType.ITIN, PiiType.EIN, PiiType.TIN,
    PiiType.EMAIL, PiiType.PHONE, PiiType.BANK_ACCOUNT, PiiType.ROUTING_NUMBER,
    PiiType.STREET, PiiType.ADDRESS, PiiType.DOB, PiiType.POLICY_NUMBER,
    PiiType.MEMBER_ID, PiiType.EMPLOYEE_ID, PiiType.DRIVERS_LICENSE,
    PiiType.PASSPORT,
}

HEADERS = ["Type", "Original value", "Pseudonym", "First seen in", "Last updated"]


def normalize(value: str) -> str:
    return " ".join(value.split()).strip(" .,;:").lower()


@dataclass
class RosterEntry:
    pii_type: PiiType
    original: str
    pseudonym: str
    first_seen_in: str = ""
    updated: str = ""

    @property
    def key(self) -> tuple[str, str]:
        return (self.pii_type.value, normalize(self.original))


@dataclass
class ClientRoster:
    """Every original-to-pseudonym pair this installation has issued."""

    path: Optional[Path] = None
    entries: dict[tuple[str, str], RosterEntry] = field(default_factory=dict)

    # -- lookup ------------------------------------------------------------

    def pseudonym_for(self, pii_type: PiiType, original: str) -> Optional[str]:
        entry = self.entries.get((pii_type.value, normalize(original)))
        return entry.pseudonym if entry else None

    def record(
        self, pii_type: PiiType, original: str, pseudonym: str, document: str = ""
    ) -> None:
        if pii_type not in ROSTER_TYPES:
            return
        key = (pii_type.value, normalize(original))
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        existing = self.entries.get(key)
        if existing:
            existing.pseudonym = pseudonym
            existing.updated = now
        else:
            self.entries[key] = RosterEntry(
                pii_type=pii_type,
                original=original,
                pseudonym=pseudonym,
                first_seen_in=document,
                updated=now,
            )

    def known_values(self, pii_type: Optional[PiiType] = None) -> list[str]:
        return [
            e.original
            for e in self.entries.values()
            if pii_type is None or e.pii_type is pii_type
        ]

    # -- persistence -------------------------------------------------------

    @classmethod
    def load(cls, path: Path) -> "ClientRoster":
        roster = cls(path=Path(path))
        if not Path(path).exists():
            return roster
        try:
            from openpyxl import load_workbook
        except ImportError:
            return roster
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return roster
        if SHEET_NAME not in workbook.sheetnames:
            workbook.close()
            return roster
        sheet = workbook[SHEET_NAME]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            try:
                pii_type = PiiType(str(row[0]))
            except ValueError:
                continue
            entry = RosterEntry(
                pii_type=pii_type,
                original=str(row[1]),
                pseudonym=str(row[2]) if len(row) > 2 and row[2] else "",
                first_seen_in=str(row[3]) if len(row) > 3 and row[3] else "",
                updated=str(row[4]) if len(row) > 4 and row[4] else "",
            )
            if entry.pseudonym:
                roster.entries[entry.key] = entry
        workbook.close()
        return roster

    def save(self, path: Optional[Path] = None) -> Path:
        target = Path(path or self.path or ROSTER_FILENAME)
        write_mapping_workbook(sorted(self.entries.values(), key=lambda e: (e.pii_type.value, e.original)), target)
        self.path = target
        return target


def write_mapping_workbook(entries: Iterable[RosterEntry], path: Path) -> Path:
    """Write the original-to-pseudonym mapping as a plain, readable workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    warning = (
        "CONFIDENTIAL - this file maps real client identifiers to their pseudonyms. "
        "Never email it, never store it with the anonymised documents."
    )
    sheet["A1"] = warning
    sheet["A1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="C00000")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    sheet.row_dimensions[1].height = 28

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")

    row = 3
    for entry in entries:
        sheet.cell(row=row, column=1, value=entry.pii_type.value)
        sheet.cell(row=row, column=2, value=entry.original)
        sheet.cell(row=row, column=3, value=entry.pseudonym)
        sheet.cell(row=row, column=4, value=entry.first_seen_in)
        sheet.cell(row=row, column=5, value=entry.updated)
        for column in range(1, len(HEADERS) + 1):
            sheet.cell(row=row, column=column).font = Font(name="Arial", size=11)
        row += 1

    widths = [22, 38, 38, 30, 18]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A3"

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return Path(path)


def mapping_path_for(output_pdf: Path, folder: Optional[Path] = None) -> Path:
    """Where the mapping for one document goes.

    Defaults to a `mappings` folder beside the output rather than next to it, so
    selecting "everything in this folder" to attach to an email does not pick it
    up.
    """
    output_pdf = Path(output_pdf)
    stem = re.sub(r"\.anonymized$", "", output_pdf.stem)
    directory = Path(folder) if folder else output_pdf.parent / "mappings"
    return directory / f"{stem}.{MAPPING_SUFFIX}"
